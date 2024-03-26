print("Student Report Creator - Created by Noah Gaunt - noah.gaunt@yahoo.co.uk")
import pandas as pd
import datetime
import os
import math
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
import subprocess
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import shutil
import pwinput
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
condFillColour = 'FCD5B4' #this changes the conditional formatting colour

thisPath = os.path.dirname(os.path.abspath(__file__))
#these headings are taken from each report downloaded
listStudentHeadings = ['Student Name',
                'Last Attendance',
                'Last Progress Check',
                'Active LPs',
                'Skills Assigned',
                'Last LP Update',
                'Last PR Sent',
                'Centre'
                ]

listAssessmentHeadings = ['Student First Name',
                          'Student Last Name',
                          'Assessment Title',
                          'Score',
                          'Date Taken'
                          ]

listActivityHeadings = ['Contact',
                        'Subject',
                        'Comments',
                        'Due Date'
                        ]

#---------#
        
print("Any personal data entered into this program is not saved.")
print("The program will take about a minute to create and open your report.")

#Initialize diffrent dates#
dtToday = datetime.datetime.today().strftime("%d.%m")
dtTodayPlusYear = datetime.datetime.today().replace(year=datetime.datetime.now().year+1).strftime("%d/%m/%Y")
dtTodayLong = datetime.datetime.today().strftime("%d_%m_%Y")
#Turn today's date into an Excel Serial number#
baseDateExcel = datetime.datetime(1900, 1, 1)
delta = datetime.datetime.today() - baseDateExcel
serialDate = str(delta.days + delta.seconds / (24*60*60))

username = input("Enter your Username: ")
password = pwinput.pwinput(prompt='Enter your Password (Will not be saved): ', mask='*')
cenNumCheck = 1
input("Please close the Student Report spreadsheet if it is open. Then press Enter and don't touch the computer until the report is open.")
start_time = time.time()

#-----------------------------------------------------------------------#

#set default download location of chrome driver
op = Options()
prefs = {"download.default_directory" : thisPath}
op.add_experimental_option("prefs",prefs)

#Open Radius
driver = webdriver.Chrome(options=op)
driver.get("*************************")

#Login to radius
driver.maximize_window()
elemUsername = driver.find_element(By.NAME,"UserName")
elemUsername.clear()
elemUsername.send_keys(username)
elemPassword = driver.find_element(By.NAME, "Password")
elemPassword.clear()
elemPassword.send_keys(password)
elemPassword.send_keys(Keys.RETURN)

#if password is incorrect, wait for user input to correct.
while driver.current_url == "*************************":
    time.sleep(0.5)
driver.get("*************************")
###Student Report Download###
#find enrollment dropdown
driver.find_element(By.XPATH, "/html/body/div[19]/div[2]/div[2]/div/div[1]/div[2]/div/span/span/span[1]").click()
time.sleep(0.5)
for i in driver.find_elements(By.XPATH,'//*[@id="enrollmentFiltersDropDownList_listbox"]//*'):
    if i.text == "Enrolment":
        i.click()
        break
driver.find_element(By.ID,"btnExport").click()
#Wait until file is downloaded
while not os.path.exists("Student Report  "+dtTodayLong+".xlsx"):
    time.sleep(0.3)
    
###Assessment Report Download###
driver.get("*************************")
driver.find_element(By.ID,"ReportStart").clear()
driver.find_element(By.ID,"ReportEnd").clear()
driver.find_element(By.ID,"btnExport").click()
    
while not os.path.exists("Assessment Report  "+dtTodayLong+".xlsx"):
    time.sleep(0.3)

###Activity Report Download###
driver.get("*************************")
driver.find_element(By.XPATH, "/html/body/div[19]/div[2]/div[3]/div/div[1]/div[1]/div[2]/div/div").click()
time.sleep(0.5)
for i in driver.find_elements(By.XPATH,'//*[@id="groupsSelectList_listbox"]//*'):
    if i.text == "Checkpoints":
        i.click()
        break
driver.find_element(By.ID,"startDateSelect").clear()
driver.find_element(By.ID,"endDateSelect").send_keys(dtTodayPlusYear)
driver.find_element(By.ID,"btnExport").click()

while not os.path.exists("My Activities Export  "+dtTodayLong+".xlsx"):
    time.sleep(0.3)
driver.close()
##-------------------------------------------------##

#Using the lists above, import required columns from both reports

rawStudents = pd.read_excel("Student Report  "+dtTodayLong+".xlsx")[listStudentHeadings].values.tolist()
rawAssessments = pd.read_excel("Assessment Report  "+dtTodayLong+".xlsx")[listAssessmentHeadings].values.tolist()
rawActivities = pd.read_excel("My Activities Export  " +dtTodayLong+".xlsx")[listActivityHeadings].values.tolist()
#Check if there are multiple centres, add centre column to spreadsheet if >1
centreCheck = pd.read_excel("Student Report  "+dtTodayLong+".xlsx")['Centre'].values.tolist()
centres = []
for c in centreCheck:
    if c not in centres:
        centres.append(c)
        if len(centres) > cenNumCheck:
            break
studentNames = []
finalExcelRows = []

#Add all gathered details to what will become the pd dataframe
for i in rawStudents:
    studentNames.append(i[0])
    if len(centres) > cenNumCheck:
        finalExcelRows.append([i[0],"","", 0,"", "", 0,i[1],i[4],i[2],i[3],i[5],i[6], "", "", "",i[7]])
    else:
        finalExcelRows.append([i[0],"","", 0,"", "", 0,i[1],i[4],i[2],i[3],i[5],i[6], "", "", ""])

#sort the assessments, only taking the most recent ext, and main level and add to the list above.
for i in range(len(rawAssessments)-1,-1,-1):
    s = studentNames.index(rawAssessments[i][0]+" "+rawAssessments[i][1])
    if "&" not in rawAssessments[i][2]:
        if "Extension" in rawAssessments[i][2]:
            if rawAssessments[i][2]=="Extension Assessment #0":
                finalExcelRows[s][4]="Extension Assessment 0"
            else:
                finalExcelRows[s][4]=rawAssessments[i][2]
            finalExcelRows[s][5]=rawAssessments[i][4]
            if math.isnan(rawAssessments[i][3]) == False:
                finalExcelRows[s][6]=rawAssessments[i][3]*100
            else:
                finalExcelRows[s][6]="N/A"
        else:
            if rawAssessments[i][2]=="Checkup Numerical Fluency - Multiplication and Division Facts":
                finalExcelRows[s][1]= "MFF"
            elif rawAssessments[i][2] == "Assessment Numerical Fluency Addition and Subtraction Facts":
                finalExcelRows[s][1] = "NF"
            elif rawAssessments[i][2]=="Assessment #0":
                finalExcelRows[s][1]="Assessment 0"
            else:
                finalExcelRows[s][1]=rawAssessments[i][2]

            finalExcelRows[s][2]=rawAssessments[i][4]
            if math.isnan(rawAssessments[i][3]) == False:
                finalExcelRows[s][3]=rawAssessments[i][3]*100
            else:
                finalExcelRows[s][3]="N/A"

#Add Open Checkpoint Activities
for i in rawActivities:
    if i[0] in studentNames:
        s = studentNames.index(i[0])
        finalExcelRows[s][13] = i[1]
        finalExcelRows[s][14] = i[2]
        finalExcelRows[s][15] = i[3]
    
#delete the reports as they are no longer needed.
try:
    os.remove(thisPath+"\Student Report  "+dtTodayLong+".xlsx")
    os.remove(thisPath+"\Assessment Report  " +dtTodayLong+".xlsx")
    os.remove(thisPath+"\My Activities Export  " +dtTodayLong+".xlsx")
except OSError as e: 
    print("Failed with:", e.strerror)
    print("Error code:", e.code)

#Open workbook and delete today's report if it exists.
if not os.path.exists(thisPath+"\Student_Report.xlsx"):
    wb = openpyxl.Workbook()
else:
    wb = openpyxl.load_workbook("Student_Report.xlsx")
if dtToday in wb.sheetnames:
    del wb[dtToday]

#create today's sheet and fill in column headers
ws = wb.create_sheet(dtToday)
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]
ws["A1"] = "Student Name"
ws["B1"] = "Last Checkup Taken"
ws["C1"] = "Last Checkup Date"
ws["D1"] = "Last Checkup Score"
ws["E1"] = "Last Ext Taken"
ws["F1"] = "Last Ext Date"
ws["G1"] = "Last Ext Score"
ws["H1"] = "Last Attendance"
ws["I1"] = "Skills Assigned"
ws["J1"] = "Last Progress Check"
ws["K1"] = "Active LPs"
ws["L1"] = "Last LP Update"
ws["M1"] = "Last PR Sent"
ws["N1"] = "Checkpoint"
ws["O1"] = "Checkpoint Comments"
ws["P1"] = "Checkpoint Date Due"
if len(centres) > cenNumCheck:
    ws["Q1"] = "Centre"

#populate report
for row in finalExcelRows:
    ws.append(row)
if len(centres) > cenNumCheck:
    tab = openpyxl.worksheet.table.Table(displayName="Report"+dtToday, ref ="A1:Q"+str(len(studentNames)+1))
else:
    tab = openpyxl.worksheet.table.Table(displayName="Report"+dtToday, ref ="A1:P"+str(len(studentNames)+1))
ws.add_table(tab)

#Conditional Formatting
fillCondFormat = PatternFill(start_color=condFillColour, end_color=condFillColour, fill_type='solid')
fillSelectedRow = PatternFill(start_color=rowHighlightColour, end_color=rowHighlightColour, fill_type='solid')
def condFormat(rng, formula, fill):
    global ws
    ws.conditional_formatting.add(rng, FormulaRule(formula = [formula], stopIfTrue = True, fill = fill))

finalRow = str(len(finalExcelRows)+1)

#Any conditional formatting goes in this dictionary in the form of {range : formula}
condFormatDict = {"H2:H"+finalRow : ["=Datevalue($H2)<"+serialDate+"-13", fillCondFormat],
                  "L2:L"+finalRow : ["=Datevalue($L2)<"+serialDate+"-13", fillCondFormat],
                  "K2:K"+finalRow : ["=$K2<>1", fillCondFormat],
                  "I2:I"+finalRow : ["=$I2<6", fillCondFormat],
                  "M2:M"+finalRow : ["=Datevalue($M2)<"+serialDate+"-31", fillCondFormat],
                  "J2:J"+finalRow : ["=IF(DATEVALUE($C2)<"+serialDate+"-90,AND(IF(ISERROR(DATEVALUE($F2)),TRUE,DATEVALUE($F2)<"+serialDate+"-90),IF(ISERROR(DATEVALUE($J2)),TRUE,DATEVALUE($J2)<"+serialDate+"-90)),FALSE)", fillCondFormat],
                  "P2:P"+finalRow : ["=$P2<"+serialDate+"", fillCondFormat]
                  }

for i in condFormatDict:
    condFormat(i, condFormatDict[i][0],condFormatDict[i][1])
wb.active = wb[dtToday]
#Code to autofit columns based on number of characters
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column_letter].width = adjusted_width

for i in range(2, int(finalRow)+1):
    ws["P"+str(i)].number_format = "[$-en-GB]d mmmm yyyy;@"
#Zoom Level
ws.sheet_view.zoomScale = 70
#Freeze Panes
ws.freeze_panes = ws['B2']

wb.save("Student_Report.xlsx")
wb.close()

#open the excel sheet
os.system("start EXCEL.EXE Student_Report.xlsx")

driver.quit()
#print("--- %s seconds ---" % (round(time.time() - start_time,0)))

